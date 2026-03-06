"""
Photo Bot — AI Photo Transformations with Credit System

A Telegram bot that applies AI-powered photo transformations using Google Gemini.
Features: credit system, promo codes, referrals, package purchases via YooMoney.
"""

import os
import io
import re
import json
import logging
import warnings
import yaml
from datetime import datetime, timedelta, timezone

from dotenv import load_dotenv
from PIL import Image
from google import genai
from google.genai import types
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaPhoto,
    ReplyKeyboardMarkup,
    LabeledPrice,
)
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    PreCheckoutQueryHandler,
    ConversationHandler,
    filters,
    ContextTypes,
)

import database as db
import notifications as notif

# ── Configuration ──────────────────────────────────────────────────────────────

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

TELEGRAM_BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
GEMINI_API_KEY = os.environ["GEMINI_API_KEY"]
YOOMONEY_PROVIDER_TOKEN = os.environ.get("YOOMONEY_PROVIDER_TOKEN", "")
ADMIN_ID = int(os.environ.get("ADMIN_ID", 0))
BOT_USERNAME = os.environ.get("BOT_USERNAME", "your_bot")
SUPPORT_USERNAME = os.environ.get("SUPPORT_USERNAME", "")  # Support account for "О проекте"

GEMINI_MODEL = "gemini-3-pro-image-preview"

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ── Load effects and categories from YAML ─────────────────────────────────────

BASE_DIR = os.path.dirname(__file__)

LOGO_PATH = None
for _ext in ("jpg", "png", "webp"):
    _p = os.path.join(BASE_DIR, "images", f"logo.{_ext}")
    if os.path.exists(_p):
        LOGO_PATH = _p
        break


def load_yaml_config() -> tuple[dict, dict]:
    """Load effects and categories from effects.yaml.

    Auto-resolves:
      - prompts/{effect_id}.txt  → effect["prompt"]
      - images/{effect_id}.jpg   → effect["example_image"]
    Categories support parent field for nesting.
    """
    yaml_path = os.path.join(BASE_DIR, "effects.yaml")
    with open(yaml_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    # Load categories (filter enabled, sort by order)
    categories = {}
    for cat_id, cat in data.get("categories", {}).items():
        if cat.get("enabled", True):
            categories[cat_id] = cat
    # Auto-resolve category images from images/{cat_id}.{ext}
    for cat_id, cat in categories.items():
        if "image" not in cat:
            for ext in ("jpg", "png", "webp"):
                img_path = os.path.join(BASE_DIR, "images", f"{cat_id}.{ext}")
                if os.path.exists(img_path):
                    cat["image"] = img_path
                    break
    categories = dict(sorted(categories.items(), key=lambda x: x[1].get("order", 999)))

    # Load effects (filter enabled, auto-resolve files, sort by order)
    effects = {}
    for effect_id, effect in data.get("effects", {}).items():
        if effect.get("enabled", True):
            # Auto-resolve prompt from prompts/{effect_id}.txt
            if "prompt" not in effect:
                prompt_path = os.path.join(BASE_DIR, "prompts", f"{effect_id}.txt")
                if os.path.exists(prompt_path):
                    with open(prompt_path, "r", encoding="utf-8") as f:
                        effect["prompt"] = f.read().strip()
                elif effect.get("type") == "free_prompt":
                    effect["prompt"] = ""   # user supplies prompt at runtime
                else:
                    logger.error(f"Prompt file not found, skipping effect: {prompt_path}")
                    continue
            # Auto-resolve example image from images/{effect_id}.jpg
            if "example_image" not in effect:
                for ext in ("jpg", "png", "webp"):
                    img_path = os.path.join(BASE_DIR, "images", f"{effect_id}.{ext}")
                    if os.path.exists(img_path):
                        effect["example_image"] = img_path
                        break
            effects[effect_id] = effect
    effects = dict(sorted(effects.items(), key=lambda x: x[1].get("order", 999)))

    return categories, effects


def get_subcategories(parent_id: str | None) -> dict:
    """Get child categories of a parent (None = top-level)."""
    return {
        k: v for k, v in CATEGORIES.items()
        if v.get("parent") == parent_id
    }


def get_effects_for(category_id: str | None) -> dict:
    """Get effects directly in a category (None = top-level / no category)."""
    return {
        k: v for k, v in TRANSFORMATIONS.items()
        if v.get("category") == category_id
    }


CATEGORIES, TRANSFORMATIONS = load_yaml_config()
logger.info(f"Loaded categories: {list(CATEGORIES.keys())}")
logger.info(f"Loaded effects: {list(TRANSFORMATIONS.keys())}")

# ── Pricing (RUB, in kopecks for Telegram) ───────────────────────────────────

PACKAGES = {
    # Note: YooKassa test mode may require min 100₽. Prices in kopecks.
    "pkg_10": {"credits": 10, "price": 9900, "label": "10 зарядов — 99 ₽"},
    "pkg_25": {"credits": 25, "price": 22900, "label": "25 зарядов — 229 ₽"},
    "pkg_50": {"credits": 50, "price": 39900, "label": "50 зарядов — 399 ₽"},
    "pkg_100": {"credits": 100, "price": 69900, "label": "100 зарядов — 699 ₽"},
}

PROMO_AMOUNTS = [10, 25, 50, 100]

# ── Conversation states ──────────────────────────────────────────────────────

(
    MAIN_MENU,
    BROWSING,
    WAITING_PHOTO,
    STORE,
    WAITING_PAYMENT,
    PROMO_INPUT,
    REFERRAL,
    ABOUT,
    ADMIN_MENU,
    ADMIN_STATS,
    ADMIN_PROMO,
    ADMIN_BULK_PROMO,
    ADMIN_REPORT,
    ADMIN_EFFECTS_REPORT,
    ADMIN_BROADCAST,
    WAITING_LUCKY_PROMPT,
    ADMIN_SOURCE_INPUT,
) = range(17)

# ── Gemini client ────────────────────────────────────────────────────────────

gemini_client = genai.Client(api_key=GEMINI_API_KEY)

# ── Helper functions ─────────────────────────────────────────────────────────


def reply_keyboard() -> ReplyKeyboardMarkup:
    """Build persistent reply keyboard (below input field)."""
    return ReplyKeyboardMarkup(
        [
            ["✨ Создать магию"],
            ["💳 Пополнить запасы", "🎁 Промокод"],
            ["👥 Пригласить друга", "ℹ️ О проекте"],
        ],
        resize_keyboard=True,
    )


def main_menu_keyboard() -> InlineKeyboardMarkup:
    """Build main menu inline keyboard."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✨ Создать магию", callback_data="menu_create")],
        [InlineKeyboardButton("💳 Пополнить запасы", callback_data="menu_store")],
        [InlineKeyboardButton("🎁 Промокод", callback_data="menu_promo")],
        [InlineKeyboardButton("👥 Пригласить друга", callback_data="menu_referral")],
        [InlineKeyboardButton("ℹ️ О проекте", callback_data="menu_about")],
    ])


def back_to_main_keyboard() -> InlineKeyboardMarkup:
    """Keyboard with just 'Назад' button."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")],
    ])


def back_to_browse_keyboard() -> InlineKeyboardMarkup:
    """Keyboard that returns user to previous browse level."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_browse")],
    ])


async def edit_message(query, text: str, reply_markup, parse_mode=None):
    """Edit message in place, preserving photo if present (for photo-bearing screens like main menu)."""
    try:
        if query.message.photo:
            await query.edit_message_caption(caption=text, reply_markup=reply_markup, parse_mode=parse_mode)
        else:
            await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
    except Exception as e:
        if "message is not modified" not in str(e).lower():
            raise


def _is_result_photo_message(message) -> bool:
    """Return True when message looks like generated result photo."""
    caption = getattr(message, "caption", None)
    return bool(caption and str(caption).startswith("✅"))


async def edit_text_screen(query, context, chat_id: int, text: str, reply_markup, parse_mode=None):
    """For text-only screens: removes stale photo if present, then edits or sends text."""
    message = getattr(query, "message", None)
    has_photo = bool(getattr(message, "photo", None))

    # If callback came from a photo message, replace it with a text message.
    if has_photo and message:
        try:
            await message.delete()
        except Exception:
            pass
        await context.bot.send_message(
            chat_id=chat_id,
            text=text,
            reply_markup=reply_markup,
            parse_mode=parse_mode,
        )
        return

    try:
        await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
    except Exception as e:
        lower = str(e).lower()
        if "message is not modified" in lower:
            return
        # Old/stale callbacks may point to messages that can't be edited anymore.
        if (
            "message to edit not found" in lower
            or "message can't be edited" in lower
            or "there is no text in the message" in lower
            or "inaccessible" in lower
            or "attribute" in lower
        ):
            await context.bot.send_message(
                chat_id=chat_id,
                text=text,
                reply_markup=reply_markup,
                parse_mode=parse_mode,
            )
            return
        raise


async def edit_main_menu_screen(query, context, chat_id: int, text: str, reply_markup):
    """Show main menu with correct logo regardless of the current message type.

    photo + logo     → edit_message_media to logo (no jump, correct image)
    photo + no logo  → edit_message_caption (no jump)
    text  + logo     → delete + send_photo (one unavoidable jump)
    text  + no logo  → edit_message_text (no jump)
    """
    try:
        if LOGO_PATH:
            if query.message.photo:
                with open(LOGO_PATH, "rb") as img:
                    await query.edit_message_media(
                        media=InputMediaPhoto(media=img, caption=text),
                        reply_markup=reply_markup,
                    )
            else:
                await query.message.delete()
                with open(LOGO_PATH, "rb") as img:
                    await context.bot.send_photo(
                        chat_id=chat_id,
                        photo=img,
                        caption=text,
                        reply_markup=reply_markup,
                    )
        else:
            if query.message.photo:
                await query.edit_message_caption(caption=text, reply_markup=reply_markup)
            else:
                await query.edit_message_text(text, reply_markup=reply_markup)
    except Exception as e:
        if "message is not modified" not in str(e).lower():
            raise


async def render_create_screen(
    context,
    chat_id: int,
    text: str,
    reply_markup,
    image_path: str | None = None,
    parse_mode: str | None = None,
) -> None:
    """Edit the single Create-flow anchor message in place.

    Reads/writes context.user_data keys:
        create_ui_message_id  — message ID of the anchor
        create_ui_is_photo    — True if the anchor is currently a photo message

    If no anchor is set, sends a new message and saves its ID.
    On any error (message gone, etc.) recreates the anchor fresh.
    """
    async def _send_new() -> None:
        if image_path and os.path.exists(image_path):
            with open(image_path, "rb") as img:
                msg = await context.bot.send_photo(
                    chat_id=chat_id,
                    photo=img,
                    caption=text,
                    reply_markup=reply_markup,
                    parse_mode=parse_mode,
                )
            context.user_data["create_ui_message_id"] = msg.message_id
            context.user_data["create_ui_is_photo"] = True
        else:
            msg = await context.bot.send_message(
                chat_id=chat_id,
                text=text or "\u200b",
                reply_markup=reply_markup,
                parse_mode=parse_mode,
            )
            context.user_data["create_ui_message_id"] = msg.message_id
            context.user_data["create_ui_is_photo"] = False

    message_id = context.user_data.get("create_ui_message_id")
    is_photo = context.user_data.get("create_ui_is_photo", False)

    if not message_id:
        await _send_new()
        return

    try:
        if image_path and os.path.exists(image_path):
            if is_photo:
                with open(image_path, "rb") as img:
                    await context.bot.edit_message_media(
                        chat_id=chat_id,
                        message_id=message_id,
                        media=InputMediaPhoto(media=img, caption=text, parse_mode=parse_mode),
                        reply_markup=reply_markup,
                    )
            else:
                # Text anchor → delete and resend as photo
                try:
                    await context.bot.delete_message(chat_id=chat_id, message_id=message_id)
                except Exception:
                    pass
                context.user_data.pop("create_ui_message_id", None)
                await _send_new()
        else:
            if is_photo:
                await context.bot.edit_message_caption(
                    chat_id=chat_id,
                    message_id=message_id,
                    caption=text,
                    reply_markup=reply_markup,
                    parse_mode=parse_mode,
                )
            else:
                await context.bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=message_id,
                    text=text or "\u200b",
                    reply_markup=reply_markup,
                    parse_mode=parse_mode,
                )
    except Exception as e:
        if "message is not modified" in str(e).lower():
            return
        # Anchor lost — recreate
        context.user_data.pop("create_ui_message_id", None)
        context.user_data.pop("create_ui_is_photo", None)
        await _send_new()


# ── Main Menu ────────────────────────────────────────────────────────────────


async def send_main_menu(bot, chat_id, text, reply_markup):
    """Send main menu with logo if available, otherwise text-only."""
    if LOGO_PATH:
        with open(LOGO_PATH, "rb") as img:
            await bot.send_photo(chat_id=chat_id, photo=img, caption=text, reply_markup=reply_markup)
    else:
        await bot.send_message(chat_id=chat_id, text=text, reply_markup=reply_markup)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle /start command. Check for referral link or deep link."""
    context.user_data.pop("lucky_photo", None)
    user = update.effective_user
    args = context.args

    # Parse deep link parameters
    referred_by = None
    auto_browse = False
    acquisition_source = None

    if args:
        param = args[0]
        # Referral link: /start ref_123456
        if param.startswith("ref_"):
            try:
                referred_by = int(param[4:])
                if referred_by == user.id:
                    referred_by = None  # Can't refer yourself
            except ValueError:
                pass
        # Browse deep link: /start browse
        elif param == "browse":
            auto_browse = True
        # Source tracking: /start src_<name>
        elif param.startswith("src_"):
            acquisition_source = param[4:]

    # Get or create user
    db_user, is_new = db.get_or_create_user(
        telegram_id=user.id,
        username=user.username,
        referred_by=referred_by,
        acquisition_source=acquisition_source,
    )

    db.update_last_active(user.id)

    credits = db_user["credits"]
    name = user.first_name or "друг"

    # If auto_browse is set, go straight to effects menu (single message)
    if auto_browse:
        context.user_data["current_category"] = None
        title, keyboard = build_browse_keyboard(None, credits)
        await update.message.reply_text(
            title,
            reply_markup=keyboard,
        )
        return BROWSING
    else:
        # Normal start - show main menu with logo
        text = f"Привет, {name}!\n⚡ Доступно зарядов: {credits}\nВыбери действие 👇"
        await send_main_menu(context.bot, update.effective_chat.id, text, reply_keyboard())
        return MAIN_MENU


async def show_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show main menu (from callback) using reply keyboard mode."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("lucky_photo", None)

    user = update.effective_user
    db_user = db.get_user(user.id)
    credits = db_user["credits"] if db_user else 0
    name = user.first_name or "друг"

    text = f"Привет, {name}!\n⚡ Доступно зарядов: {credits}\nВыбери действие 👇"

    # Keep generated result photos, but remove other callback-origin messages
    # so main menu stays visually consistent as a fresh reply-keyboard screen.
    is_result_photo = _is_result_photo_message(getattr(query, "message", None))
    if not is_result_photo:
        try:
            await query.message.delete()
        except Exception:
            pass

    await send_main_menu(context.bot, update.effective_chat.id, text, reply_keyboard())
    return MAIN_MENU


async def restart_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle restart button (callback version of /start) in reply keyboard mode."""
    query = update.callback_query
    await query.answer()

    user = update.effective_user
    db_user = db.get_user(user.id)
    credits = db_user["credits"] if db_user else 0
    name = user.first_name or "друг"

    # Clear any stale user data
    context.user_data.clear()

    text = f"Привет, {name}!\n⚡ Доступно зарядов: {credits}\nВыбери действие 👇"

    is_result_photo = _is_result_photo_message(getattr(query, "message", None))
    if not is_result_photo:
        try:
            await query.message.delete()
        except Exception:
            pass

    await send_main_menu(context.bot, update.effective_chat.id, text, reply_keyboard())
    return MAIN_MENU


async def back_to_browse(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Return to previous browse category."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("lucky_photo", None)

    user = update.effective_user
    db_user = db.get_user(user.id)
    credits = db_user["credits"] if db_user else 0

    category_id = context.user_data.get("previous_category")
    context.user_data["current_category"] = category_id

    title, keyboard = build_browse_keyboard(category_id, credits)

    cat_image = CATEGORIES.get(category_id or "", {}).get("image") if category_id else None
    image_path = (cat_image if cat_image and os.path.exists(cat_image) else None) or LOGO_PATH

    await render_create_screen(context, user.id, title, keyboard, image_path)
    return BROWSING


# ── Reply Keyboard Handlers ─────────────────────────────────────────────────


def build_browse_keyboard(category_id: str | None, credits: int) -> tuple[str, InlineKeyboardMarkup]:
    """Build keyboard showing subcategories + effects for a category.

    category_id=None → top-level (the "Create" screen).
    Returns (title_text, keyboard).
    """
    buttons = []

    # Effects at this level
    for eff_id, eff in get_effects_for(category_id).items():
        buttons.append([InlineKeyboardButton(eff["label"], callback_data=f"effect_{eff_id}")])

    # Subcategories
    for cat_id, cat in get_subcategories(category_id).items():
        buttons.append([InlineKeyboardButton(cat["label"], callback_data=f"cat_{cat_id}")])

    # Back button
    if category_id is None:
        buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")])
    else:
        parent = CATEGORIES.get(category_id, {}).get("parent")
        back_data = f"cat_{parent}" if parent else "browse_root"
        buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data=back_data)])

    title = f"⚡ Доступно зарядов: {credits}" if category_id is None else ""
    return title, InlineKeyboardMarkup(buttons)


async def handle_reply_create(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle '✨ Создать магию' from reply keyboard."""
    user = update.effective_user
    db_user = db.get_user(user.id)
    credits = db_user["credits"] if db_user else 0

    context.user_data['current_category'] = None
    # Clear stale anchor so render_create_screen sends a fresh message
    context.user_data.pop("create_ui_message_id", None)
    context.user_data.pop("create_ui_is_photo", None)
    context.user_data.pop("lucky_photo", None)

    title, keyboard = build_browse_keyboard(None, credits)
    await render_create_screen(context, update.effective_chat.id, title, keyboard, LOGO_PATH)
    return BROWSING


async def handle_reply_store(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle '💳 Пополнить запасы' from reply keyboard."""
    context.user_data.pop("lucky_photo", None)
    buttons = [
        [InlineKeyboardButton(pkg["label"], callback_data=f"buy_{key}")]
        for key, pkg in PACKAGES.items()
    ]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")])
    await update.message.reply_text("Выбери пакет:", reply_markup=InlineKeyboardMarkup(buttons))
    return STORE


async def handle_reply_promo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle '🎁 Промокод' from reply keyboard."""
    context.user_data.pop("lucky_photo", None)
    await update.message.reply_text(
        "Введи промокод:",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")],
        ]),
    )
    return PROMO_INPUT


async def handle_reply_referral(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle '👥 Пригласить друга' from reply keyboard."""
    context.user_data.pop("lucky_photo", None)
    user = update.effective_user
    ref_link = f"https://t.me/{BOT_USERNAME}?start=ref_{user.id}"
    await update.message.reply_text(
        f"Приглашай друзей и получай\n+3 заряда за каждого!\n\nТвоя ссылка:\n{ref_link}",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")],
        ]),
    )
    return REFERRAL


# ── Create Magic Flow ────────────────────────────────────────────────────────


async def show_browse_root(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show top-level browse screen (from inline menu)."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("lucky_photo", None)
    user = update.effective_user
    db_user = db.get_user(user.id)
    credits = db_user["credits"] if db_user else 0

    context.user_data['current_category'] = None
    # Adopt the current callback message as Create anchor when available.
    if getattr(query, "message", None):
        context.user_data["create_ui_message_id"] = query.message.message_id
        context.user_data["create_ui_is_photo"] = bool(getattr(query.message, "photo", None))

    title, keyboard = build_browse_keyboard(None, credits)
    await render_create_screen(context, update.effective_chat.id, title, keyboard, LOGO_PATH)
    return BROWSING


async def browse_category(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Navigate into a category/subcategory — generic handler for any depth."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("lucky_photo", None)
    user = update.effective_user
    db_user = db.get_user(user.id)
    credits = db_user["credits"] if db_user else 0

    category_id = query.data.removeprefix("cat_")

    # Validate category exists
    if category_id and category_id not in CATEGORIES:
        error_text = "❌ Категория не найдена\n\nНажми кнопку ниже, чтобы начать заново:"
        error_keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Начать заново", callback_data="restart")],
        ])
        await edit_text_screen(query, context, update.effective_chat.id, error_text, error_keyboard)
        return MAIN_MENU

    context.user_data['current_category'] = category_id

    title, keyboard = build_browse_keyboard(category_id, credits)

    cat_image = CATEGORIES.get(category_id, {}).get("image")
    image_path = cat_image if cat_image and os.path.exists(cat_image) else None

    await render_create_screen(context, update.effective_chat.id, title, keyboard, image_path)
    return BROWSING


async def select_effect(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """User selected an effect. Check credits and show description."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("lucky_photo", None)

    effect_id = query.data.replace("effect_", "")
    if effect_id not in TRANSFORMATIONS:
        await query.edit_message_text("Неизвестный эффект.", reply_markup=back_to_main_keyboard())
        return MAIN_MENU

    user = update.effective_user
    db_user = db.get_user(user.id)
    credits = db_user["credits"] if db_user else 0

    # Check credits
    if credits < 1:
        # Store category before showing error
        current_category = context.user_data.get('current_category')
        back_callback = f"cat_{current_category}" if current_category else "browse_root"

        # Credits exhausted message (inline UI)
        message = (
            "😮‍💨 Заряды кончились. Бывает.\n\n"
            "Но останавливаться необязательно:\n\n"
            "💳 Пополнить → от 99 ₽\n"
            "👥 Позвать друга → +3 заряда бесплатно"
        )

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("💳 Пополнить", callback_data="menu_store")],
            [InlineKeyboardButton("👥 Позвать друга", callback_data="menu_referral")],
            [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
        ])

        await render_create_screen(
            context, update.effective_chat.id, message, keyboard, parse_mode="HTML"
        )
        return BROWSING

    # Store selected effect and remember which category we came from
    context.user_data["effect_id"] = effect_id
    context.user_data["previous_category"] = context.user_data.get('current_category')

    effect = TRANSFORMATIONS[effect_id]

    # Build back button that returns to the category we came from
    previous_category = context.user_data.get("previous_category")
    back_callback = f"cat_{previous_category}" if previous_category else "browse_root"
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
    ])

    tips = (effect.get('tips') or '').strip()
    best_input = (effect.get('best_input') or '').strip()
    parts = []
    if tips:
        parts.append(tips)
    if best_input:
        parts.append(f"📷 Лучше всего подойдёт: {best_input}")
    parts.append("Отправь мне фото для обработки 👇")
    message = "\n\n".join(parts)

    example_image = effect.get("example_image")
    image_path = example_image if example_image and os.path.exists(example_image) else None
    await render_create_screen(context, update.effective_chat.id, message, keyboard, image_path)
    return WAITING_PHOTO


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Receive photo and process it."""
    effect_id = context.user_data.get("effect_id")
    if not effect_id or effect_id not in TRANSFORMATIONS:
        # Session lost - offer restart
        await update.message.reply_text(
            "❌ Сессия истекла\n\nНажми кнопку ниже, чтобы начать заново:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Начать заново", callback_data="restart")],
            ]),
        )
        return MAIN_MENU

    user = update.effective_user

    # Deduct credit
    if not db.deduct_credit(user.id):
        # Credits exhausted message (inline UI)
        message = (
            "😮‍💨 Заряды кончились. Бывает.\n\n"
            "Но останавливаться необязательно:\n\n"
            "💳 Пополнить → от 99 ₽\n"
            "👥 Позвать друга → +3 заряда бесплатно"
        )

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("💳 Пополнить", callback_data="menu_store")],
            [InlineKeyboardButton("👥 Позвать друга", callback_data="menu_referral")],
            [InlineKeyboardButton("⬅️ Главное меню", callback_data="back_to_main")],
        ])

        await update.message.reply_text(
            message,
            reply_markup=keyboard,
            parse_mode="HTML",
        )

        return MAIN_MENU

    # Download photo
    photo_file = await update.message.photo[-1].get_file()
    photo_bytes = await photo_file.download_as_bytearray()

    effect = TRANSFORMATIONS[effect_id]

    # free_prompt branch: store photo, ask for user's text prompt (before spinner + try/finally)
    if effect.get("type") == "free_prompt":
        db.refund_credit(user.id)  # charge later, when user actually submits their text
        try:
            buf = io.BytesIO()
            Image.open(io.BytesIO(bytes(photo_bytes))).save(buf, format="PNG")
            context.user_data["lucky_photo"] = buf.getvalue()
        except Exception as e:
            logger.error("Failed to process photo for free_prompt: %s", e)
            # Credit already refunded above — do NOT call add_credits again
            await update.message.reply_text("❌ Не удалось обработать фото. Попробуй ещё раз.")
            return WAITING_PHOTO

        previous_category = context.user_data.get("previous_category")
        back_callback = f"cat_{previous_category}" if previous_category else "browse_root"
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
        ])
        # Delete the old effect-selection anchor so its buttons can't be tapped,
        # then reply to the photo (puts prompt request below it) and adopt as new anchor.
        old_anchor_id = context.user_data.pop("create_ui_message_id", None)
        context.user_data.pop("create_ui_is_photo", None)
        if old_anchor_id:
            try:
                await context.bot.delete_message(
                    chat_id=update.effective_chat.id, message_id=old_anchor_id
                )
            except Exception:
                pass
        msg = await update.message.reply_text(
            "✅ Фото получено!\n\nТеперь напиши свой PROMPT 👇",
            reply_markup=keyboard,
        )
        context.user_data["create_ui_message_id"] = msg.message_id
        context.user_data["create_ui_is_photo"] = False
        return WAITING_LUCKY_PROMPT

    status_msg = await update.message.reply_text("⏳ Создаю магию...")

    try:
        input_image = Image.open(io.BytesIO(bytes(photo_bytes)))

        # Call Gemini
        logger.info(f"Calling Gemini model: {GEMINI_MODEL}")
        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[effect["prompt"], input_image],
            config=types.GenerateContentConfig(
                response_modalities=["Text", "Image"],
            ),
        )
        # Log model info from response if available
        if hasattr(response, 'model_version'):
            logger.info(f"Gemini response model_version: {response.model_version}")
        logger.info(f"Gemini response candidates: {len(response.candidates) if response.candidates else 0}")

        # Extract result image
        result_image = None
        result_text = None
        for part in response.parts:
            if part.inline_data is not None:
                # Convert Gemini response to PIL Image
                image_data = part.inline_data.data
                result_image = Image.open(io.BytesIO(image_data))
            elif part.text is not None:
                result_text = part.text

        if result_image is None:
            # Record failed generation, then refund credit
            db.record_generation(user.id, effect_id, status="failed")
            new_balance = db.refund_credit(user.id)
            msg = f"❌ Что-то пошло не так\n\nКредит возвращён на баланс.\n⚡ Доступно зарядов: {new_balance}"
            if result_text:
                msg += f"\n\nОтвет модели: {result_text[:200]}"

            # Build back button that returns to the category we came from
            previous_category = context.user_data.get("previous_category")
            back_callback = f"cat_{previous_category}" if previous_category else "browse_root"

            keyboard = InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Попробовать снова", callback_data=f"effect_{effect_id}")],
                [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
            ])
            await status_msg.edit_text(msg, reply_markup=keyboard)
            return BROWSING

        # Record generation for statistics
        db.record_generation(user.id, effect_id)
        db.update_last_active(user.id)

        # Credit referrer on first generation (only for referrer's first 10 referrals)
        referrer_id = db.credit_referral_on_generation(user.id)
        if referrer_id:
            db.add_credits(referrer_id, 3)
            logger.info(f"Credited referrer {referrer_id} with 3 credits (generation) for user {user.id}")

        # Get updated balance
        db_user = db.get_user(user.id)
        remaining = db_user["credits"] if db_user else 0

        # Real-time notification triggers
        gen_count = db.get_user_generation_count(user.id)

        # N2: Credits Running Low
        if remaining == 1:
            await notif.send_credits_low_warning(user.id)


        # Send result with navigation buttons
        previous_category = context.user_data.get("previous_category")
        context.user_data["current_category"] = previous_category
        back_callback = f"cat_{previous_category}" if previous_category else "browse_root"
        result_keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Попробовать снова", callback_data=f"effect_{effect_id}")],
            [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
        ])

        output_buffer = io.BytesIO()
        result_image.save(output_buffer, format="PNG")
        output_buffer.seek(0)

        await status_msg.delete()
        await update.message.reply_photo(
            photo=output_buffer,
            caption=f"✅ {effect['label']}\n⚡ Осталось зарядов: {remaining}",
            reply_markup=result_keyboard,
        )

        # Delete old anchor (replaced by result photo buttons)
        old_anchor_id = context.user_data.pop("create_ui_message_id", None)
        context.user_data.pop("create_ui_is_photo", None)
        if old_anchor_id:
            try:
                await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=old_anchor_id)
            except Exception:
                pass

    except Exception as e:
        logger.error("Error during transformation: %s", e, exc_info=True)
        # Record failed generation, then refund credit
        db.record_generation(user.id, effect_id, status="failed")
        new_balance = db.refund_credit(user.id)

        # Build back button that returns to the category we came from
        previous_category = context.user_data.get("previous_category")
        back_callback = f"cat_{previous_category}" if previous_category else "browse_root"

        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Попробовать снова", callback_data=f"effect_{effect_id}")],
            [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
        ])
        await status_msg.edit_text(
            f"❌ Что-то пошло не так\n\nКредит возвращён на баланс.\n⚡ Доступно зарядов: {new_balance}\n\nОшибка: {str(e)[:100]}",
            reply_markup=keyboard,
        )
        return BROWSING
    finally:
        context.user_data.pop("effect_id", None)

    return BROWSING


async def photo_expected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle non-photo message when photo is expected."""
    # Build back button that returns to the category we came from
    previous_category = context.user_data.get("previous_category")
    back_callback = f"cat_{previous_category}" if previous_category else "browse_root"

    await update.message.reply_text(
        "📸 Сначала фото — потом магия!",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
        ]),
    )
    return WAITING_PHOTO


async def handle_lucky_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Receive user's text prompt and apply it to the stored photo."""
    effect_id = context.user_data.get("effect_id")
    photo_bytes = context.user_data.get("lucky_photo")

    if not effect_id or not photo_bytes or effect_id not in TRANSFORMATIONS:
        await update.message.reply_text(
            "❌ Сессия истекла\n\nНажми кнопку ниже, чтобы начать заново:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Начать заново", callback_data="restart")],
            ]),
        )
        return MAIN_MENU

    user = update.effective_user
    user_text = update.message.text.strip()

    if not user_text:
        await update.message.reply_text("✏️ Пустой запрос не считается 🙈 Напиши что-нибудь!")
        return WAITING_LUCKY_PROMPT

    if not db.deduct_credit(user.id):
        context.user_data.pop("lucky_photo", None)
        context.user_data.pop("effect_id", None)
        message = (
            "😮‍💨 Заряды кончились. Бывает.\n\n"
            "Но останавливаться необязательно:\n\n"
            "💳 Пополнить → от 99 ₽\n"
            "👥 Позвать друга → +3 заряда бесплатно"
        )
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("💳 Пополнить", callback_data="menu_store")],
            [InlineKeyboardButton("👥 Позвать друга", callback_data="menu_referral")],
            [InlineKeyboardButton("⬅️ Главное меню", callback_data="back_to_main")],
        ])
        await update.message.reply_text(message, reply_markup=keyboard, parse_mode="HTML")
        return MAIN_MENU

    effect = TRANSFORMATIONS[effect_id]
    status_msg = await update.message.reply_text("⏳ Создаю магию...")

    try:
        input_image = Image.open(io.BytesIO(photo_bytes))

        logger.info(f"Calling Gemini model (free_prompt): {GEMINI_MODEL}")
        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[user_text, input_image],
            config=types.GenerateContentConfig(
                response_modalities=["Text", "Image"],
            ),
        )

        result_image = None
        result_text = None
        for part in response.parts:
            if part.inline_data is not None:
                result_image = Image.open(io.BytesIO(part.inline_data.data))
            elif part.text is not None:
                result_text = part.text

        if result_image is None:
            db.record_generation(user.id, effect_id, status="failed")
            new_balance = db.refund_credit(user.id)
            previous_category = context.user_data.get("previous_category")
            back_callback = f"cat_{previous_category}" if previous_category else "browse_root"
            keyboard = InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Попробовать снова", callback_data=f"effect_{effect_id}")],
                [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
            ])
            await status_msg.edit_text(
                f"❌ Что-то пошло не так\n\nКредит возвращён.\n⚡ Доступно зарядов: {new_balance}",
                reply_markup=keyboard,
            )
            return BROWSING

        db.record_generation(user.id, effect_id)
        db.update_last_active(user.id)
        referrer_id = db.credit_referral_on_generation(user.id)
        if referrer_id:
            db.add_credits(referrer_id, 3)
            logger.info(f"Credited referrer {referrer_id} with 3 credits (generation) for user {user.id}")

        db_user = db.get_user(user.id)
        remaining = db_user["credits"] if db_user else 0
        if remaining == 1:
            await notif.send_credits_low_warning(user.id)

        output_buffer = io.BytesIO()
        result_image.save(output_buffer, format="PNG")
        output_buffer.seek(0)

        result_keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Попробовать снова", callback_data=f"effect_{effect_id}")],
            [InlineKeyboardButton("⬅️ Назад", callback_data="browse_root")],
        ])

        await status_msg.delete()
        await update.message.reply_photo(
            photo=output_buffer,
            caption=f"✅ {effect['label']}\n⚡ Осталось зарядов: {remaining}",
            reply_markup=result_keyboard,
        )

        # Delete old anchor (replaced by result photo buttons)
        old_anchor_id = context.user_data.pop("create_ui_message_id", None)
        context.user_data.pop("create_ui_is_photo", None)
        if old_anchor_id:
            try:
                await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=old_anchor_id)
            except Exception:
                pass

    except Exception as e:
        logger.error("Error during free_prompt generation: %s", e, exc_info=True)
        db.record_generation(user.id, effect_id, status="failed")
        new_balance = db.refund_credit(user.id)
        previous_category = context.user_data.get("previous_category")
        back_callback = f"cat_{previous_category}" if previous_category else "browse_root"
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Попробовать снова", callback_data=f"effect_{effect_id}")],
            [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
        ])
        await status_msg.edit_text(
            f"❌ Что-то пошло не так\n\nКредит возвращён.\n⚡ Доступно зарядов: {new_balance}\n\nОшибка: {str(e)[:100]}",
            reply_markup=keyboard,
        )
        return BROWSING
    finally:
        context.user_data.pop("effect_id", None)
        context.user_data.pop("lucky_photo", None)

    return BROWSING


async def lucky_prompt_expected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle non-text message when prompt text is expected."""
    previous_category = context.user_data.get("previous_category")
    back_callback = f"cat_{previous_category}" if previous_category else "browse_root"
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("⬅️ Назад", callback_data=back_callback)],
    ])
    await render_create_screen(
        context,
        update.effective_chat.id,
        "Мне нужен текст, а не это 😄 Напиши словами что сделать с фото",
        keyboard,
    )
    return WAITING_LUCKY_PROMPT


# ── Store Flow ───────────────────────────────────────────────────────────────


async def show_store(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show package store."""
    query = update.callback_query
    await query.answer()

    buttons = [
        [InlineKeyboardButton(pkg["label"], callback_data=f"buy_{key}")]
        for key, pkg in PACKAGES.items()
    ]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")])

    text = "Выбери пакет:"
    keyboard = InlineKeyboardMarkup(buttons)

    await edit_text_screen(query, context, update.effective_chat.id, text, keyboard)
    return STORE


async def buy_package(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Send payment invoice for selected package."""
    query = update.callback_query
    await query.answer()

    package_id = query.data.replace("buy_", "")
    if package_id not in PACKAGES:
        await query.edit_message_text("Неизвестный пакет.", reply_markup=back_to_main_keyboard())
        return MAIN_MENU

    package = PACKAGES[package_id]
    context.user_data["pending_package"] = package_id

    try:
        # Delete the menu message
        await query.delete_message()

        # Prepare fiscal receipt data (required by YooKassa for Russian tax law)
        # NOTE: Invoice price is in kopecks, but receipt amount must be in RUBLES
        price_rubles = package["price"] / 100

        receipt_data = {
            "receipt": {
                "items": [
                    {
                        "description": f"{package['credits']} зарядов",
                        "quantity": 1,
                        "amount": {
                            "value": price_rubles,
                            "currency": "RUB"
                        },
                        "vat_code": 1,  # No VAT
                        "payment_mode": "full_payment",
                        "payment_subject": "service"
                    }
                ]
            }
        }

        # Send invoice with fiscal data
        # Email will be collected on payment form (required for receipt delivery)
        invoice_msg = await context.bot.send_invoice(
            chat_id=update.effective_chat.id,
            title=f"Пакет {package['credits']} зарядов",
            description=f"Пополнение баланса на {package['credits']} зарядов",
            payload=f"package_{package_id}_{update.effective_user.id}",
            currency="RUB",
            prices=[LabeledPrice(f"{package['credits']} зарядов", package["price"])],
            provider_token=YOOMONEY_PROVIDER_TOKEN,
            provider_data=json.dumps(receipt_data),
            need_email=True,
            send_email_to_provider=True,
        )
        context.user_data["pending_invoice_message_id"] = invoice_msg.message_id
        db.record_invoice(update.effective_user.id, package_id)

        # Send cancel button separately (invoices can't have inline buttons)
        cancel_msg = await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="Нажми кнопку ниже, чтобы отменить покупку:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Назад", callback_data="cancel_payment")]
            ]),
        )
        context.user_data["pending_cancel_message_id"] = cancel_msg.message_id
        return WAITING_PAYMENT
    except Exception as e:
        logger.error(f"Payment error: {e}")
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"❌ Ошибка платежа: {e}\n\nПопробуйте позже.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ К выбору пакетов", callback_data="menu_store")],
            ]),
        )
        return STORE


async def cancel_payment(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel payment and return to store."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop("pending_package", None)
    db.mark_invoice_cancelled(update.effective_user.id)

    # Remove cancel prompt message
    try:
        await query.message.delete()
    except Exception:
        pass

    # Remove invoice message so chat doesn't keep stale payment UI
    invoice_message_id = context.user_data.pop("pending_invoice_message_id", None)
    if invoice_message_id:
        try:
            await context.bot.delete_message(
                chat_id=update.effective_chat.id,
                message_id=invoice_message_id,
            )
        except Exception:
            pass

    context.user_data.pop("pending_cancel_message_id", None)

    buttons = [
        [InlineKeyboardButton(pkg["label"], callback_data=f"buy_{key}")]
        for key, pkg in PACKAGES.items()
    ]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")])
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text="Выбери пакет:",
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return STORE


async def show_main_menu_fresh(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Send a fresh main menu message (not edit)."""
    user = update.effective_user
    db_user = db.get_user(user.id)
    credits = db_user["credits"] if db_user else 0
    name = user.first_name or "друг"
    text = f"Привет, {name}!\n⚡ Доступно зарядов: {credits}\nВыбери действие 👇"
    await send_main_menu(context.bot, user.id, text, main_menu_keyboard())
    return MAIN_MENU


async def recover_stale_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Recover from stale inline callbacks after deploy/restart.
    Triggered only when callback was not handled by ConversationHandler.
    """
    query = update.callback_query
    if not query:
        return

    # Acknowledge click quickly so user doesn't see a hanging spinner.
    await query.answer("Бот обновился, открываю меню...")

    user = update.effective_user
    db_user = db.get_user(user.id)
    credits = db_user["credits"] if db_user else 0
    name = user.first_name or "друг"
    text = f"Привет, {name}!\n⚡ Доступно зарядов: {credits}\nВыбери действие 👇"

    # Never delete a generated result photo (caption always starts with ✅).
    is_result_photo = _is_result_photo_message(getattr(query, "message", None))
    if not is_result_photo:
        try:
            await query.message.delete()
        except Exception:
            pass

    await send_main_menu(context.bot, user.id, text, reply_keyboard())


async def pre_checkout(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Approve the payment."""
    query = update.pre_checkout_query
    await query.answer(ok=True)


async def successful_payment(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle successful payment."""
    user = update.effective_user
    package_id = context.user_data.pop("pending_package", None)

    if package_id and package_id in PACKAGES:
        package = PACKAGES[package_id]
        credits = package["credits"]
        price_rub = package["price"] // 100  # Convert kopecks to rubles

        # Add credits and record purchase
        new_balance = db.add_credits(user.id, credits)
        db.record_purchase(user.id, credits, price_rub)
        db.update_last_active(user.id)
        db.mark_invoice_paid(user.id, package_id)

        # Credit referrer on first payment (for referrer's 11th+ referrals)
        referrer_id = db.credit_referral_on_payment(user.id)
        if referrer_id:
            db.add_credits(referrer_id, 3)
            logger.info(f"Credited referrer {referrer_id} with 3 credits (payment) for user {user.id}")

        # N7: First Purchase Thank You
        if db.get_user_purchase_count(user.id) == 1:
            await notif.send_first_purchase_thanks(user.id)

        await update.message.reply_text(
            f"✅ Оплата прошла!\n+{credits} зарядов добавлено\n\n⚡ Доступно зарядов: {new_balance}",
            reply_markup=back_to_main_keyboard(),
        )
    else:
        await update.message.reply_text(
            "Оплата получена, но произошла ошибка. Свяжитесь с поддержкой.",
            reply_markup=back_to_main_keyboard(),
        )

    return MAIN_MENU


# ── Promo Code Flow ──────────────────────────────────────────────────────────


async def show_promo_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Ask for promo code."""
    query = update.callback_query
    await query.answer()

    text = "Введи промокод:"
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")],
    ])

    await edit_text_screen(query, context, update.effective_chat.id, text, keyboard)
    return PROMO_INPUT


async def handle_promo_code(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle promo code input."""
    user = update.effective_user
    code = update.message.text.strip()

    success, message, credits = db.redeem_promo_code(user.id, code)

    if success:
        db.update_last_active(user.id)
        db_user = db.get_user(user.id)
        new_balance = db_user["credits"] if db_user else 0
        await update.message.reply_text(
            f"✅ Промокод активирован!\n+{credits} зарядов добавлено\n\n⚡ Доступно зарядов: {new_balance}",
            reply_markup=back_to_main_keyboard(),
        )
    else:
        await update.message.reply_text(
            f"❌ {message}",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Попробовать другой", callback_data="menu_promo")],
                [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")],
            ]),
        )

    return MAIN_MENU


# ── Referral Flow ────────────────────────────────────────────────────────────


async def show_referral(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show referral info and link."""
    query = update.callback_query
    await query.answer()

    user = update.effective_user
    ref_link = f"https://t.me/{BOT_USERNAME}?start=ref_{user.id}"

    text = f"Приглашай друзей и получай\n+3 заряда за каждого!\n\nТвоя ссылка:\n{ref_link}"
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")],
    ])

    await edit_text_screen(query, context, update.effective_chat.id, text, keyboard)
    return REFERRAL


# ── About Flow ──────────────────────────────────────────────────────────────


async def show_about(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show about/disclaimer info."""
    query = update.callback_query
    await query.answer()

    text = (
        "ℹ️ О проекте\n\n"
        "Проект предназначен для людей достигших возраста 18+ "
        "и демонстрирует работу нейросетей.\n\n"
        "Все созданные изображения не могут быть использованы "
        "для рекламных или иных порочащих репутацию других граждан целей.\n\n"
        "Если вам кажется что ваши права нарушены или у вас возникли "
        "вопросы/предложения по работе проекта – пишите в нашу поддержку."
    )

    buttons = []
    if SUPPORT_USERNAME:
        buttons.append([InlineKeyboardButton("✉️ Написать в поддержку", url=f"https://t.me/{SUPPORT_USERNAME}")])
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")])

    keyboard = InlineKeyboardMarkup(buttons)

    await edit_text_screen(query, context, update.effective_chat.id, text, keyboard)
    return ABOUT


async def show_about_from_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show about/disclaimer info (from reply keyboard text)."""
    context.user_data.pop("lucky_photo", None)
    text = (
        "ℹ️ О проекте\n\n"
        "Проект предназначен для людей достигших возраста 18+ "
        "и демонстрирует работу нейросетей.\n\n"
        "Все созданные изображения не могут быть использованы "
        "для рекламных или иных порочащих репутацию других граждан целей.\n\n"
        "Если вам кажется что ваши права нарушены или у вас возникли "
        "вопросы/предложения по работе проекта – пишите в нашу поддержку."
    )

    buttons = []
    if SUPPORT_USERNAME:
        buttons.append([InlineKeyboardButton("✉️ Написать в поддержку", url=f"https://t.me/{SUPPORT_USERNAME}")])
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")])

    await update.message.reply_text(
        text,
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return ABOUT


# ── Admin Menu ───────────────────────────────────────────────────────────────


async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle /admin command."""
    context.user_data.pop("lucky_photo", None)
    user = update.effective_user

    logger.info(f"Admin attempt: user.id={user.id}, ADMIN_ID={ADMIN_ID}")

    if user.id != ADMIN_ID:
        await update.message.reply_text(f"Доступ запрещён. (Your ID: {user.id})")
        return ConversationHandler.END

    await update.message.reply_text(
        "🔐 Админ-панель",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 Статистика", callback_data="admin_stats")],
            [InlineKeyboardButton("📈 Weekly Report", callback_data="admin_report")],
            [InlineKeyboardButton("🗂 Raw Data", callback_data="admin_effects_report")],
            [InlineKeyboardButton("🎁 Подарочный промокод", callback_data="admin_promo")],
            [InlineKeyboardButton("🎟 Массовый промокод", callback_data="admin_bulk_promo")],
            [InlineKeyboardButton("📢 Рассылка новых эффектов", callback_data="admin_broadcast")],
            [InlineKeyboardButton("🔗 Source Links", callback_data="admin_source_links")],
            [InlineKeyboardButton("🏠 Выход", callback_data="back_to_main")],
        ]),
    )
    return ADMIN_MENU


async def show_admin_stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show bot statistics."""
    query = update.callback_query
    await query.answer()

    stats = db.get_stats()

    # Build effect stats text
    effect_lines = []
    for effect_id, effect in TRANSFORMATIONS.items():
        count = stats["effect_stats"].get(effect_id, 0)
        # Get emoji from label
        emoji = effect["label"].split()[0]
        name = " ".join(effect["label"].split()[1:])
        effect_lines.append(f"{emoji} {name}: {count}")

    effects_text = "\n".join(effect_lines)

    # Build package stats text
    package_lines = []
    package_stats = stats.get("package_stats", {})
    for pkg_id, pkg in PACKAGES.items():
        credits = pkg["credits"]
        pkg_data = package_stats.get(credits, {"count": 0, "revenue": 0})
        package_lines.append(f"{credits} зарядов: {pkg_data['count']} шт. ({pkg_data['revenue']} ₽)")

    packages_text = "\n".join(package_lines)

    text = (
        f"📊 Статистика бота\n\n"
        f"Пользователей: {stats['total_users']}\n"
        f"Всего генераций: {stats['total_generations']}\n"
        f"Куплено пакетов: {stats['total_purchases']}\n"
        f"Доход: {stats['total_revenue']} ₽\n\n"
        f"── По эффектам ──\n{effects_text}\n\n"
        f"── По пакетам ──\n{packages_text}"
    )

    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
        ]),
    )
    return ADMIN_STATS


async def show_admin_promo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show promo code creation menu."""
    query = update.callback_query
    await query.answer()

    buttons = [
        [InlineKeyboardButton(f"{amount} зарядов", callback_data=f"create_promo_{amount}")]
        for amount in PROMO_AMOUNTS
    ]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")])

    await query.edit_message_text(
        "🎁 Создать промокод\n\nСколько зарядов даёт промокод?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return ADMIN_PROMO


async def create_promo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Create a promo code."""
    query = update.callback_query
    await query.answer()

    amount = int(query.data.replace("create_promo_", ""))
    code = db.create_promo_code(credits=amount, max_uses=1)

    await query.edit_message_text(
        f"✅ Промокод создан!\n\nКод: {code}\nДаёт: +{amount} зарядов",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🎁 Создать ещё", callback_data="admin_promo")],
            [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
        ]),
    )
    return ADMIN_MENU


BULK_USE_OPTIONS = [5, 10, 20, 50, 100]
BULK_EXPIRY_OPTIONS = [3, 7, 14, 30]  # days


async def show_admin_bulk_promo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Step 1: choose credits for bulk promo."""
    query = update.callback_query
    await query.answer()

    buttons = [
        [InlineKeyboardButton(f"{amount} зарядов", callback_data=f"bulk_credits_{amount}")]
        for amount in PROMO_AMOUNTS
    ]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")])

    await query.edit_message_text(
        "🎟 Массовый промокод\n\nШаг 1/3 — Сколько зарядов даёт промокод?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return ADMIN_BULK_PROMO


async def bulk_select_credits(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Step 2: choose max uses."""
    query = update.callback_query
    await query.answer()

    credits = int(query.data.replace("bulk_credits_", ""))
    context.user_data["bulk_credits"] = credits

    buttons = [
        [InlineKeyboardButton(f"{n} раз", callback_data=f"bulk_uses_{n}")]
        for n in BULK_USE_OPTIONS
    ]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="admin_bulk_promo")])

    await query.edit_message_text(
        f"🎟 Массовый промокод\n\nЗарядов: {credits}\nШаг 2/3 — Сколько раз можно использовать?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return ADMIN_BULK_PROMO


async def bulk_select_uses(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Step 3: choose expiry."""
    query = update.callback_query
    await query.answer()

    uses = int(query.data.replace("bulk_uses_", ""))
    context.user_data["bulk_uses"] = uses

    credits = context.user_data.get("bulk_credits", "?")
    buttons = [
        [InlineKeyboardButton(f"{days} дней", callback_data=f"bulk_expiry_{days}")]
        for days in BULK_EXPIRY_OPTIONS
    ]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="admin_bulk_promo")])

    await query.edit_message_text(
        f"🎟 Массовый промокод\n\nЗарядов: {credits} · Использований: {uses}\nШаг 3/3 — Срок действия?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return ADMIN_BULK_PROMO


async def bulk_select_expiry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Create bulk promo code and show result."""
    query = update.callback_query
    await query.answer()

    days = int(query.data.replace("bulk_expiry_", ""))
    credits = context.user_data.get("bulk_credits", 10)
    uses = context.user_data.get("bulk_uses", 10)
    expires_at = datetime.now(timezone.utc) + timedelta(days=days)

    code = db.create_promo_code(credits=credits, max_uses=uses, expires_at=expires_at)

    await query.edit_message_text(
        f"✅ Массовый промокод создан!\n\n"
        f"Код: {code}\n"
        f"Зарядов: +{credits}\n"
        f"Макс. использований: {uses}\n"
        f"Истекает: {expires_at.strftime('%d.%m.%Y')}",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🎟 Создать ещё", callback_data="admin_bulk_promo")],
            [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
        ]),
    )
    context.user_data.pop("bulk_credits", None)
    context.user_data.pop("bulk_uses", None)
    return ADMIN_MENU


async def admin_back(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Go back to admin menu."""
    query = update.callback_query
    await query.answer()

    await query.edit_message_text(
        "🔐 Админ-панель",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 Статистика", callback_data="admin_stats")],
            [InlineKeyboardButton("📈 Weekly Report", callback_data="admin_report")],
            [InlineKeyboardButton("🗂 Raw Data", callback_data="admin_effects_report")],
            [InlineKeyboardButton("🎁 Подарочный промокод", callback_data="admin_promo")],
            [InlineKeyboardButton("🎟 Массовый промокод", callback_data="admin_bulk_promo")],
            [InlineKeyboardButton("📢 Рассылка новых эффектов", callback_data="admin_broadcast")],
            [InlineKeyboardButton("🔗 Source Links", callback_data="admin_source_links")],
            [InlineKeyboardButton("🏠 Выход", callback_data="back_to_main")],
        ]),
    )
    return ADMIN_MENU


async def show_admin_source_links(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show list of source tracking links."""
    query = update.callback_query
    await query.answer()
    names = db.get_source_links()
    if names:
        lines = "\n".join(f"• https://t.me/{BOT_USERNAME}?start=src_{n}" for n in names)
    else:
        lines = "No source links yet."
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Create new", callback_data="admin_source_create")],
        [InlineKeyboardButton("⬅️ Back", callback_data="admin_back")],
    ])
    await query.edit_message_text(f"🔗 Source Links\n\n{lines}", reply_markup=keyboard)
    return ADMIN_MENU


async def show_admin_source_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Prompt admin to enter a source name."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("Enter source name (letters/digits only, e.g. TG):")
    return ADMIN_SOURCE_INPUT


async def handle_admin_source_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Save new source link and show the generated URL."""
    name = update.message.text.strip()
    if not re.match(r'^[A-Za-z0-9]{1,32}$', name):
        await update.message.reply_text("Invalid name. Use ASCII letters and digits only, max 32 chars:")
        return ADMIN_SOURCE_INPUT
    created = db.create_source_link(name)
    if created:
        link = f"https://t.me/{BOT_USERNAME}?start=src_{name}"
        await update.message.reply_text(f"Created:\n{link}")
    else:
        await update.message.reply_text(f"Source '{name}' already exists.")
    await update.message.reply_text(
        "🔐 Админ-панель",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 Статистика", callback_data="admin_stats")],
            [InlineKeyboardButton("📈 Weekly Report", callback_data="admin_report")],
            [InlineKeyboardButton("🗂 Raw Data", callback_data="admin_effects_report")],
            [InlineKeyboardButton("🎁 Подарочный промокод", callback_data="admin_promo")],
            [InlineKeyboardButton("🎟 Массовый промокод", callback_data="admin_bulk_promo")],
            [InlineKeyboardButton("📢 Рассылка новых эффектов", callback_data="admin_broadcast")],
            [InlineKeyboardButton("🔗 Source Links", callback_data="admin_source_links")],
            [InlineKeyboardButton("🏠 Выход", callback_data="back_to_main")],
        ]),
    )
    return ADMIN_MENU


async def show_admin_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Send weekly core metrics report as Telegram text."""
    query = update.callback_query
    await query.answer()

    r = db.get_weekly_report()

    def pct(val: float) -> str:
        return f"{round(val * 100)}%"

    text = (
        "📈 Weekly Report\n\n"
        "ACQUISITION\n"
        f"New Users (7d): {r['new_users']}{db._wow(r['new_users'], r['new_users_prior'])}\n"
        f"Activation Rate (24h): {pct(r['activation_rate'])}{db._wow(r['activation_rate'], r['activation_rate_prior'])}\n\n"
        "ENGAGEMENT\n"
        f"Returning Active Users (7d): {r['returning_active']}{db._wow(r['returning_active'], r['returning_active_prior'])}\n"
        f"Week-1 Activity Rate: {pct(r['week1_activity_rate'])} (cohort: {r['cohort_size']}){db._wow(r['week1_activity_rate'], r['week1_activity_rate_prior'])}\n"
        f"Avg Generations / Active User: {r['avg_gens']:.1f}{db._wow(r['avg_gens'], r['avg_gens_prior'])}\n\n"
        "MONETIZATION\n"
        f"Week-1 Payment Rate: {pct(r['week1_payment_rate'])} (cohort: {r['cohort_size']}){db._wow(r['week1_payment_rate'], r['week1_payment_rate_prior'])}\n"
        f"Revenue (7d): {r['revenue_7d']} ₽{db._wow(r['revenue_7d'], r['revenue_prior'])}\n"
        f"RPAU-7d: {r['rpau_7d']:.1f} ₽{db._wow(r['rpau_7d'], r['rpau_prior'])}"
    )

    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
        ]),
    )
    return ADMIN_REPORT


async def show_admin_effects_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show format selection for raw data export."""
    query = update.callback_query
    await query.answer()

    await query.edit_message_text(
        "🗂 Raw Data — выбери формат:",
        reply_markup=InlineKeyboardMarkup([
            [
                InlineKeyboardButton("📊 XLSX", callback_data="admin_effects_report_xlsx"),
                InlineKeyboardButton("📄 CSV (zip)", callback_data="admin_effects_report_csv"),
            ],
            [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
        ]),
    )
    return ADMIN_EFFECTS_REPORT


async def show_admin_effects_report_xlsx(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Send raw data (users, generations, purchases) as Excel file."""
    import io
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import Font

    query = update.callback_query
    await query.answer()
    await query.edit_message_text("⏳ Generating XLSX export...")

    conn = db.get_connection()
    try:
        c = conn.cursor()
        wb = Workbook()

        for table in ["users", "generations", "purchases"]:
            c.execute(f"SELECT * FROM {table}")
            rows = c.fetchall()
            cols = [d[0] for d in c.description]

            ws = wb.create_sheet(title=table.capitalize())
            ws.append(cols)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in rows:
                ws.append(list(row))

        del wb["Sheet"]  # remove default empty sheet

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=buf,
            filename=f"raw_data_{date.today()}.xlsx",
        )

        await query.edit_message_text(
            "✅ XLSX sent! (Users / Generations / Purchases)",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
            ]),
        )
    except Exception as e:
        await query.edit_message_text(
            f"❌ Export failed: {e}",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
            ]),
        )
    finally:
        conn.close()

    return ADMIN_EFFECTS_REPORT


async def show_admin_effects_report_csv(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Send raw data (users, generations, purchases) as CSV files in a ZIP archive."""
    import io
    import csv
    import zipfile
    from datetime import date

    query = update.callback_query
    await query.answer()
    await query.edit_message_text("⏳ Generating CSV export...")

    conn = db.get_connection()
    try:
        c = conn.cursor()

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for table in ["users", "generations", "purchases"]:
                c.execute(f"SELECT * FROM {table}")
                rows = c.fetchall()
                cols = [d[0] for d in c.description]

                # utf-8-sig writes a BOM so Excel auto-detects encoding correctly
                csv_buf = io.StringIO()
                csv_buf.write("\ufeff")
                writer = csv.writer(csv_buf)
                writer.writerow(cols)
                writer.writerows(rows)

                zf.writestr(f"{table}.csv", csv_buf.getvalue().encode("utf-8"))

        zip_buf.seek(0)

        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=zip_buf,
            filename=f"raw_data_{date.today()}.zip",
        )

        await query.edit_message_text(
            "✅ CSV sent! (users.csv / generations.csv / purchases.csv)",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
            ]),
        )
    except Exception as e:
        await query.edit_message_text(
            f"❌ Export failed: {e}",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
            ]),
        )
    finally:
        conn.close()

    return ADMIN_EFFECTS_REPORT


# ── N5: Admin Broadcast (New Effects) ─────────────────────────────────────────


async def show_admin_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """N5: Ask admin for the effects list to broadcast."""
    query = update.callback_query
    await query.answer()

    # Count eligible users
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT COUNT(*) FROM users WHERE last_active_at >= datetime('now', '-14 days')"
    )
    count = cursor.fetchone()[0]
    conn.close()

    await query.edit_message_text(
        f"📢 Рассылка новых эффектов (N5)\n\n"
        f"Получателей: {count} (активные за 14 дней)\n\n"
        "Отправь список новых эффектов (каждый с новой строки).\n"
        "Например:\n"
        "🎭 Гангстер из 90-х\n"
        "👑 Египетский фараон",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Назад", callback_data="admin_back")],
        ]),
    )
    return ADMIN_BROADCAST


async def handle_admin_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """N5: Send new effects broadcast to active users."""
    import asyncio

    effects_list = update.message.text.strip()

    # Get active users
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT telegram_id, credits FROM users WHERE last_active_at >= datetime('now', '-14 days')"
    )
    users = cursor.fetchall()
    conn.close()

    await update.message.reply_text(f"⏳ Отправляю {len(users)} пользователям...")

    sent_count = 0
    for u in users:
        success = await notif.send_new_effects(u["telegram_id"], effects_list)
        if success:
            sent_count += 1
        await asyncio.sleep(0.1)  # Rate limit

    await update.message.reply_text(
        f"✅ Рассылка завершена!\nОтправлено: {sent_count}/{len(users)}",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Админ-панель", callback_data="admin_back")],
        ]),
    )
    return ADMIN_MENU


# ── Main ─────────────────────────────────────────────────────────────────────


def main() -> None:
    """Start the bot."""
    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()

    # Initialize notification system
    notif.init_notifications(app.bot)

    # Reply keyboard handlers — shared across all user-facing states
    reply_kb = [
        MessageHandler(filters.Regex("^✨ Создать магию$"), handle_reply_create),
        MessageHandler(filters.Regex("^💳 Пополнить запасы$"), handle_reply_store),
        MessageHandler(filters.Regex("^🎁 Промокод$"), handle_reply_promo),
        MessageHandler(filters.Regex("^👥 Пригласить друга$"), handle_reply_referral),
        MessageHandler(filters.Regex("^ℹ️ О проекте$"), show_about_from_text),
    ]

    # per_message=False is intentional: allows old inline buttons to work after restarts
    warnings.filterwarnings("ignore", message="If 'per_message=False'", category=UserWarning)
    # Main conversation handler
    conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            CommandHandler("admin", admin_command),
            # Allow old inline buttons to work even after bot restarts.
            CallbackQueryHandler(restart_bot, pattern="^restart$"),
            CallbackQueryHandler(back_to_browse, pattern="^back_to_browse$"),
            CallbackQueryHandler(show_browse_root, pattern="^(menu_create|browse_root)$"),
            CallbackQueryHandler(show_store, pattern="^menu_store$"),
            CallbackQueryHandler(show_promo_input, pattern="^menu_promo$"),
            CallbackQueryHandler(show_referral, pattern="^menu_referral$"),
            CallbackQueryHandler(show_about, pattern="^menu_about$"),
            CallbackQueryHandler(show_main_menu, pattern="^(back_to_main|main_menu)$"),
            CallbackQueryHandler(browse_category, pattern="^cat_"),
            CallbackQueryHandler(select_effect, pattern="^effect_"),
            CallbackQueryHandler(buy_package, pattern="^buy_"),
            CallbackQueryHandler(cancel_payment, pattern="^cancel_payment$"),
            # Allow reply keyboard text actions to recover after restart too.
            *reply_kb,
        ],
        states={
            MAIN_MENU: [
                # Inline keyboard handlers
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(back_to_browse, pattern="^back_to_browse$"),
                CallbackQueryHandler(show_browse_root, pattern="^menu_create$"),
                CallbackQueryHandler(show_store, pattern="^menu_store$"),
                CallbackQueryHandler(show_promo_input, pattern="^menu_promo$"),
                CallbackQueryHandler(show_referral, pattern="^menu_referral$"),
                CallbackQueryHandler(show_about, pattern="^menu_about$"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
                # Effect retry
                CallbackQueryHandler(select_effect, pattern="^effect_"),
            ] + reply_kb,
            BROWSING: [
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(show_browse_root, pattern="^browse_root$"),
                CallbackQueryHandler(browse_category, pattern="^cat_"),
                CallbackQueryHandler(select_effect, pattern="^effect_"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
            ] + reply_kb,
            WAITING_PHOTO: [
                MessageHandler(filters.PHOTO, handle_photo),
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(back_to_browse, pattern="^back_to_browse$"),
                CallbackQueryHandler(show_browse_root, pattern="^browse_root$"),
                CallbackQueryHandler(browse_category, pattern="^cat_"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
            ] + reply_kb + [
                MessageHandler(~filters.PHOTO & ~filters.COMMAND, photo_expected),
            ],
            STORE: [
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(show_store, pattern="^menu_store$"),
                CallbackQueryHandler(buy_package, pattern="^buy_"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
            ] + reply_kb,
            WAITING_PAYMENT: [
                MessageHandler(filters.SUCCESSFUL_PAYMENT, successful_payment),
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(show_store, pattern="^menu_store$"),
                CallbackQueryHandler(cancel_payment, pattern="^cancel_payment$"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
            ] + reply_kb,
            PROMO_INPUT: reply_kb + [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_promo_code),
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
            ],
            WAITING_LUCKY_PROMPT: reply_kb + [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_lucky_prompt),
                MessageHandler(~filters.TEXT & ~filters.COMMAND, lucky_prompt_expected),
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(back_to_browse, pattern="^back_to_browse$"),
                CallbackQueryHandler(show_browse_root, pattern="^browse_root$"),
                CallbackQueryHandler(browse_category, pattern="^cat_"),
                CallbackQueryHandler(select_effect, pattern="^effect_"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
            ],
            REFERRAL: [
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
            ] + reply_kb,
            ABOUT: [
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
            ] + reply_kb,
            ADMIN_MENU: [
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(show_admin_stats, pattern="^admin_stats$"),
                CallbackQueryHandler(show_admin_report, pattern="^admin_report$"),
                CallbackQueryHandler(show_admin_effects_report, pattern="^admin_effects_report$"),
                CallbackQueryHandler(show_admin_effects_report_xlsx, pattern="^admin_effects_report_xlsx$"),
                CallbackQueryHandler(show_admin_effects_report_csv, pattern="^admin_effects_report_csv$"),
                CallbackQueryHandler(show_admin_promo, pattern="^admin_promo$"),
                CallbackQueryHandler(show_admin_bulk_promo, pattern="^admin_bulk_promo$"),
                CallbackQueryHandler(show_admin_broadcast, pattern="^admin_broadcast$"),
                CallbackQueryHandler(show_admin_source_links, pattern="^admin_source_links$"),
                CallbackQueryHandler(show_admin_source_input, pattern="^admin_source_create$"),
                CallbackQueryHandler(admin_back, pattern="^admin_back$"),
                CallbackQueryHandler(show_main_menu, pattern="^back_to_main$"),
            ],
            ADMIN_STATS: [
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(admin_back, pattern="^admin_back$"),
            ],
            ADMIN_REPORT: [
                CallbackQueryHandler(admin_back, pattern="^admin_back$"),
            ],
            ADMIN_EFFECTS_REPORT: [
                CallbackQueryHandler(show_admin_effects_report_xlsx, pattern="^admin_effects_report_xlsx$"),
                CallbackQueryHandler(show_admin_effects_report_csv, pattern="^admin_effects_report_csv$"),
                CallbackQueryHandler(admin_back, pattern="^admin_back$"),
            ],
            ADMIN_PROMO: [
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(create_promo, pattern="^create_promo_"),
                CallbackQueryHandler(admin_back, pattern="^admin_back$"),
            ],
            ADMIN_BULK_PROMO: [
                CallbackQueryHandler(restart_bot, pattern="^restart$"),
                CallbackQueryHandler(show_admin_bulk_promo, pattern="^admin_bulk_promo$"),
                CallbackQueryHandler(bulk_select_credits, pattern="^bulk_credits_"),
                CallbackQueryHandler(bulk_select_uses, pattern="^bulk_uses_"),
                CallbackQueryHandler(bulk_select_expiry, pattern="^bulk_expiry_"),
                CallbackQueryHandler(admin_back, pattern="^admin_back$"),
            ],
            ADMIN_BROADCAST: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_admin_broadcast),
                CallbackQueryHandler(admin_back, pattern="^admin_back$"),
            ],
            ADMIN_SOURCE_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_admin_source_name),
                CallbackQueryHandler(admin_back, pattern="^admin_back$"),
            ],
        },
        fallbacks=[
            CommandHandler("start", start),
            CommandHandler("admin", admin_command),
        ],
    )

    app.add_handler(conv_handler)
    # Fallback recovery for stale/unknown callback_data after deploys.
    app.add_handler(CallbackQueryHandler(recover_stale_callback))

    # PreCheckoutQueryHandler must be at app level
    app.add_handler(PreCheckoutQueryHandler(pre_checkout))

    logger.info("Photo bot started. Polling...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
